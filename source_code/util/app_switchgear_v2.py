"""
Switchgear Health Index Streamlit Application
----------------------------------------------
Predict health condition of switchgear units (A1–A31)
using LSTM Autoencoder and visualize fleet status
"""

import pandas as pd
import numpy as np
import streamlit as st
import os
import glob
import csv
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from io import BytesIO
from datetime import datetime

from source_code.util.feature_engineering import  SwitchgearFeatureEngineering
from source_code.util.data_processing import prepare_switchgear_df
from source_code.util.generate_output import plot_signal_with_daily_risk, plot_daily_monitor
from source_code.util.prediction import (load_model_bundle, detect_anomaly,
                                         classify_risk, classify_risk2, risk_7day,
                                         daily_risk, prepare_result_dataset)
from source_code.util.db import (
    get_engine, ensure_table_exists, upsert_dataframe, reset_table, fetch_all_results,
    test_connection
)

def count_files(folder_path):
    """Count number of files in a folder"""
    if not os.path.exists(folder_path):
        return 0

    return len([
        f for f in os.listdir(folder_path)
        if os.path.isfile(os.path.join(folder_path, f))
    ])
def clear_folder(folder_path):
    """Delete all files in a folder"""
    if os.path.exists(folder_path):
        for f in os.listdir(folder_path):
            file_path = os.path.join(folder_path, f)
            if os.path.isfile(file_path):
                os.remove(file_path)

def status_explain(window_days):

    with st.expander("ℹ️ How Switchgear Status (Safe / Medium / High) is Determined"):

        st.markdown("### Status Meaning")

        legend_df = pd.DataFrame({
            "Status": ["🟢 Safe", "🟡 Medium Risk", "🔴 High Risk", "⚪ No Data"],
            "Meaning": [
                "Switchgear operating normally with no significant abnormal signals",
                "Warning condition — multiple medium or high risk signals detected",
                "Critical condition — high-risk signals persist continuously",
                "No prediction available — system cannot determine machine condition"
            ]
        })

        st.table(legend_df)

        # -----------------------------------------
        # Evaluation window
        # -----------------------------------------

        if window_days == 7:
            window_text = "most recent **7 days**"
            window_value = "7 days"
        else:
            window_text = "most recent **1 day**"
            window_value = "1 day"

        st.markdown("### Evaluation Window")

        st.info(
            f"""
            The **machine status is determined using the {window_text} of prediction results**.

            Each day the model produces a **Health Index** value:

            • **1 → Normal signal**  
            • **2 → Medium risk signal**  
            • **3 → High risk signal**

            The system evaluates how frequently abnormal signals occur within the **latest {window_value}**.
            """
        )

        st.markdown("### Decision Logic")

        logic_df = pd.DataFrame({
            "Indicator": [
                "High-risk events",
                "Medium-risk events",
                "Combined abnormal events"
            ],
            "Calculation": [
                f"Count of days with health_index = 3 within last {window_value}",
                f"Count of days with health_index = 2 within last {window_value}",
                f"Count of days with health_index = 2 or 3 within last {window_value}"
            ],
            "Threshold": [
                "7 days" if window_days == 7 else ">= 1 day",
                "Included in combined rule",
                "> 4 days" if window_days == 7 else ">= 1 day"
            ],
            "Impact": [
                "Triggers High Risk",
                "Contributes to Medium Risk",
                "Triggers Medium Risk"
            ]
        })

        st.table(logic_df)

        st.markdown("### Final Decision Rule")

        if window_days == 7:
            decision_df = pd.DataFrame({
                "Condition": [
                    "health_index = 3 for all 7 days",
                    "Number of (health_index = 2 or 3) > 4 days",
                    "Otherwise"
                ],
                "Final Status": [
                    "🔴 High Risk",
                    "🟡 Medium Risk",
                    "🟢 Safe"
                ]
            })
        else:
            decision_df = pd.DataFrame({
                "Condition": [
                    "Health Index = 3",
                    "Health Index = 2",
                    "Health Index = 1"
                ],
                "Final Status": [
                    "🔴 High Risk",
                    "🟡 Medium Risk",
                    "🟢 Safe"
                ]
            })

        st.table(decision_df)

        st.info(
            f"""
            In summary, the system analyzes the **recent {window_value} behavior of the switchgear**.
            Persistent abnormal signals increase the risk level of the equipment.
            """
        )

def compute_machine_status(df, mode="7D"):

    # ----------------------
    # validation
    # ----------------------
    if df is None or df.empty:
        return 0

    if "health_index" not in df.columns:
        return 0

    df = df.copy()

    df["health_index"] = pd.to_numeric(df["health_index"], errors="coerce")

    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")

    df = df.dropna(subset=["health_index"])

    if df.empty:
        return 0

    # ----------------------
    # convert hourly → daily
    # ----------------------
    if "date" in df.columns:

        df["day"] = df["date"].dt.date

        df_daily = (
            df.groupby("day")["health_index"]
            .last()
            .reset_index()
            .sort_values("day")
        )

    else:
        df_daily = df.copy()

    # ----------------------
    # Mode 1D
    # ----------------------
    if mode == "1D":

        last_val = df_daily["health_index"].iloc[-1]

        if pd.isna(last_val):
            return 0

        return int(last_val)

    # ----------------------
    # Mode 7D
    # ----------------------
    df_last = df_daily.tail(7)

    if df_last.empty:
        return 0

    n_red = (df_last["health_index"] == 3).sum()
    n_medium = (df_last["health_index"] == 2).sum()

    # ----------------------
    # decision logic
    # ----------------------
    if n_red == 7:
        return 3

    elif (n_red + n_medium) > 4:
        return 2

    else:
        return 1
def status_risk(path_output="data/output"):

    st.header("Switchgear Fleet Status")

    use_7day = st.toggle(
        "Use 7-Day Evaluation Window",
        value=True
    )

    if use_7day:
        status_mode = "7D"
        window_days = 7
        st.caption("Status calculated from the most recent 7 days of Health Index predictions.")
    else:
        status_mode = "1D"
        window_days = 1
        st.caption("Status calculated from the most recent 1 day Health Index prediction.")

    machines = [f"A{i}" for i in range(1, 32) if i not in [16, 17]]

    cols = st.columns(5)

    for i, mc in enumerate(machines):

        path = f"{path_output}/{mc}.csv"
        status = 0

        if os.path.exists(path):

            try:
                df = pd.read_csv(path)
            except pd.errors.EmptyDataError:
                df = pd.DataFrame()

            if not df.empty:
                status = compute_machine_status(df, status_mode)

        col = cols[i % 5]

        with col:

            st.markdown(f"**{mc}**")

            if status == 1:
                st.success("🟢 Safe")

            elif status == 2:
                st.warning("🟡 Medium Risk")

            elif status == 3:
                st.error("🔴 High Risk")

            else:
                st.info("⚪ No Data")

    return window_days


def display_risk_by_case(path_raw="data/raw_data",path_output="data/output"):

    st.header("Switchgear Health Index")

    # -------------------------------------------------
    # Select mode first
    # -------------------------------------------------

    mode = st.radio(
        "Select Dataset Type",
        ["Switchgear A1–A31"]
    )

    # =================================================
    # MODE 1 : NORMAL SWITCHGEAR
    # =================================================

    if mode == "Switchgear A1–A31":

        switchgears = [f"A{i}" for i in range(1, 32) if i not in [16, 17]]

        switchgear = st.selectbox(
            "Select Switchgear",
            switchgears
        )

        path = f"{path_output}/{switchgear}.csv"

        if not os.path.exists(path):
            st.warning("No prediction file found")
            return

        try:
            df_output = pd.read_csv(path)
        except pd.errors.EmptyDataError:
            st.warning("Prediction file empty")
            return

        if df_output.empty:
            st.warning("No prediction data")
            return

        

        df_result = pd.read_csv(f"{path_output}/{switchgear}.csv")
        plot_signal_with_daily_risk(df_result,signal_type="pd")
        plot_signal_with_daily_risk(df_result,signal_type="temp")
        plot_daily_monitor(df_result)

        # st.dataframe(df_result.head())


def plot_feature(df):

    # -----------------------------
    # Load dataset
    # -----------------------------
    #dataset_path = f"data/raw_data/{machine}.csv"

    #df = pd.read_csv(dataset_path)

    # -----------------------------
    # Convert datetime
    # -----------------------------
    df["datetime"] = pd.to_datetime(df["Start"], errors="coerce")

    # -----------------------------
    # Convert numeric
    # -----------------------------
    exclude_cols = ["Start", "End", "datetime"]

    for col in df.columns:
        if col not in exclude_cols:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.sort_values("datetime")

    # -----------------------------
    # Feature groups
    # -----------------------------
    temp_upper = [
        "UPPER_BUS_PHASE_A",
        "UPPER_BUS_PHASE_B",
        "UPPER_BUS_PHASE_C"
    ]

    temp_lower = [
        "LOWER_BUS_PHASE_A",
        "LOWER_BUS_PHASE_B",
        "LOWER_BUS_PHASE_C"
    ]

    temp_outgoing = [
        "OUTGOING_PHASE_A",
        "OUTGOING_PHASE_B",
        "OUTGOING_PHASE_C"
    ]

    pd_cols = [
        "UPPER_BUS_PD",
        "LOWER_BUS_PD",
        "SPOUT_PD",
        "OUTGOING_PD"
    ]

    elec_cols = [
        "POWER",
        "VOLTAGE",
        "FREQUENCY"
    ]

    # -----------------------------
    # Plot function
    # -----------------------------
    def plot_group(features, title):

        fig, ax = plt.subplots(figsize=(14,4))

        for col in features:
            if col in df.columns:
                ax.plot(
                    df["datetime"],
                    df[col],
                    label=col,
                    linewidth=2
                )

        ax.set_title(title)
        ax.set_xlabel("Datetime")
        ax.set_ylabel("Value")
        ax.legend()
        ax.grid(alpha=0.3)

        plt.tight_layout()

        st.pyplot(fig)
        plt.close()

    # -----------------------------
    # Temperature
    # -----------------------------
    st.subheader("Temperature Sensors")

    plot_group(temp_upper, "Upper Bus Temperature")
    plot_group(temp_lower, "Lower Bus Temperature")
    plot_group(temp_outgoing, "Outgoing Temperature")

    # -----------------------------
    # PD
    # -----------------------------
    st.subheader("Partial Discharge")

    plot_group(pd_cols, "PD Sensors")

    # -----------------------------
    # Electrical
    # -----------------------------
    st.subheader("Electrical")

    plot_group(elec_cols, "Electrical Signals")

def run_prediction_for_machine(
    machine,
    model,
    scaler,
    feature_names,
    threshold,
    window_size,
    q1,
    q2
):

    dataset_path = f"data/raw_data/{machine}.csv"
    save_path = f"data/output/{machine}.csv"

    os.makedirs("data/output", exist_ok=True)

    empty_df = pd.DataFrame(columns=["date", "health_index", "Status"])

    # -----------------------------
    # dataset check
    # -----------------------------
    if not os.path.exists(dataset_path):

        empty_df.to_csv(save_path, index=False)

        return machine, "No Data", "Dataset file not found"

    try:
        df = pd.read_csv(dataset_path)

    except Exception as e:

        empty_df.to_csv(save_path, index=False)

        return machine, "Fail", f"CSV read error: {e}"

    # -----------------------------
    # preprocessing
    # -----------------------------
    try:

        df_raw = prepare_switchgear_df(df)

        fe = SwitchgearFeatureEngineering(df_raw)

        df_hourly, df_daily, df_feature = fe.run()

    except Exception as e:

        empty_df.to_csv(save_path, index=False)

        return machine, "Fail", f"Preprocess error: {e}"

    if df_daily is None or df_daily.empty:

        empty_df.to_csv(save_path, index=False)

        return machine, "No Data", "No daily data after preprocessing"

    # -----------------------------
    # feature validation
    # -----------------------------
    missing = [c for c in feature_names if c not in df_feature.columns]

    if missing:

        empty_df.to_csv(save_path, index=False)

        return machine, "Fail", f"Missing features: {missing}"

    df_feature = df_feature.dropna(subset=feature_names)

    if df_feature.empty:

        empty_df.to_csv(save_path, index=False)

        return machine, "No Data", "All rows removed due to NaN features"

    # -----------------------------
    # anomaly detection
    # -----------------------------
    try:

        df_feature = detect_anomaly(
            df_feature,
            feature_names,
            scaler,
            model,
            threshold,
            window_size=window_size
        )

    except Exception as e:

        empty_df.to_csv(save_path, index=False)

        return machine, "Fail", f"Anomaly detection error: {e}"

    # -----------------------------
    # merge daily features
    # -----------------------------
    try:

        cols_to_add = [c for c in df_daily.columns if c not in df_feature.columns]

        cols_to_add = ["date"] + [c for c in cols_to_add if c != "date"]

        df_daily = df_feature.merge(
            df_daily[cols_to_add],
            on="date",
            how="left"
        )

        # -----------------------------
        # ensure datetime exists
        # -----------------------------
        if "datetime" not in df_daily.columns:

            if "date" in df_daily.columns:
                df_daily["datetime"] = pd.to_datetime(df_daily["date"])

            elif "Start" in df_daily.columns:
                df_daily["datetime"] = pd.to_datetime(df_daily["Start"])

    except Exception as e:

        empty_df.to_csv(save_path, index=False)

        return machine, "Fail", f"Merge error: {e}"

    # -----------------------------
    # risk classification
    # -----------------------------
    try:

        sw_daily = classify_risk(df_daily, q1, q2)

        daily_sw = daily_risk(sw_daily)

        daily_sw = risk_7day(daily_sw)

        daily_sw["health_index"] = np.select(
        [
            daily_sw["anomaly_score"] <= q1,
            daily_sw["anomaly_score"] <= q2
        ],
        [
            1,
            1
        ],
        default=2
    )

        daily_sw = classify_risk2(daily_sw)

    except Exception as e:

        empty_df.to_csv(save_path, index=False)

        return machine, "Fail", f"Risk classification error: {e}"

    # -----------------------------
    # prepare output
    # -----------------------------
    try:

        df_result = prepare_result_dataset(df_hourly, daily_sw)

        if df_result is None or df_result.empty:

            empty_df.to_csv(save_path, index=False)

            return machine, "No Data", "Postprocess produced empty output"

        status = compute_machine_status(df_result)

        df_result["Status"] = status

    except Exception as e:

        empty_df.to_csv(save_path, index=False)

        return machine, "Fail", f"Postprocess error: {e}"

    # -----------------------------
    # save
    # -----------------------------
    df_result.to_csv(save_path, index=False)

    return machine, "Success", f"Status = {status}"

def run_fleet_prediction():

    machines = [f"A{i}" for i in range(1, 32)]

    results = []

    progress = st.progress(0)

    # -----------------------------
    # LOAD MODEL ONCE
    # -----------------------------
    model_path = r"source_code/model_bundle_switchgear"
    model_name = "lstm_autoencoder_switchgear.keras"

    model, scaler, config, meta = load_model_bundle(
        path=model_path,
        model_name=model_name
    )

    feature_names = meta["features"]
    threshold = meta["high_th"]
    window_size = meta["window_size"]

    q1 = meta["high_th"]
    q2 = meta["thresholds"]["daily_max"]

    # -----------------------------
    # RUN FLEET
    # -----------------------------
    for i, mc in enumerate(machines):

        machine, status, message = run_prediction_for_machine(
            mc,
            model,
            scaler,
            feature_names,
            threshold,
            window_size,
            q1,
            q2
        )

        results.append({
            "Switchgear": machine,
            "Run Status": status,
            "Detail": message
        })

        progress.progress((i + 1) / len(machines))

    df_result = pd.DataFrame(results)

    #st.success("Fleet prediction completed")

    #st.subheader("Prediction Execution Summary")

    with st.expander("Execution Summary", expanded=False):

        st.dataframe(
            df_result,
            width='stretch'
            # use_container_width=True
#             `use_container_width` will be removed after 2025-12-31.
#             For `use_container_width=True`, use `width='stretch'`. For `use_container_width=False`, use `width='content'`.
        )

    return df_result

# =====================================================
# SAVE TO DATABASE / DOWNLOAD RESULTS
# =====================================================
SWITCHGEAR_OUTPUT_FOLDER = "data/output"


def _list_output_files(output_folder=SWITCHGEAR_OUTPUT_FOLDER):
    return sorted(glob.glob(f"{output_folder}/*.csv"))


def _machine_name_from_path(csv_path):
    return os.path.splitext(os.path.basename(csv_path))[0]


def _build_results_summary(output_folder=SWITCHGEAR_OUTPUT_FOLDER):
    """Per-machine last-row summary, read straight from disk (no DB involved)."""

    rows = []

    for csv_path in _list_output_files(output_folder):
        machine = _machine_name_from_path(csv_path)

        try:
            df = pd.read_csv(csv_path)
            last = df.iloc[-1] if not df.empty else None
        except Exception as e:
            rows.append({"Switchgear": machine, "Rows": 0, "Last Date": None,
                         "Health Index": None, "Status": f"Read error: {e}"})
            continue

        rows.append({
            "Switchgear": machine,
            "Rows": len(df),
            "Last Date": last.get("date") if last is not None else None,
            "Health Index": last.get("health_index") if last is not None else None,
            "Status": last.get("Status") if last is not None else None,
        })

    return pd.DataFrame(rows)


def _build_workbook_from_output_folder(output_folder=SWITCHGEAR_OUTPUT_FOLDER):
    # write_only streams rows straight to a temp file instead of keeping
    # every cell as an in-memory object across all sheets, which matters
    # here: ~31 sheets x thousands of rows in normal mode can be a large
    # memory spike on a memory-constrained deployment.
    wb = openpyxl.Workbook(write_only=True)

    for csv_path in _list_output_files(output_folder):
        ws = wb.create_sheet(title=_machine_name_from_path(csv_path))
        with open(csv_path) as f_input:
            for row in csv.reader(f_input):
                ws.append(row)

    return wb


def _build_workbook_from_dataframe(df: pd.DataFrame):
    wb = openpyxl.Workbook(write_only=True)

    for machine, group in df.groupby("machine"):
        ws = wb.create_sheet(title=str(machine))
        ws.append(list(group.columns))
        for row in group.itertuples(index=False):
            ws.append(list(row))

    return wb


def _upsert_output_folder(engine, output_folder=SWITCHGEAR_OUTPUT_FOLDER):
    ensure_table_exists(engine)

    files = _list_output_files(output_folder)
    total_rows = 0
    progress = st.progress(0.0)

    # One transaction PER MACHINE, not one for the whole loop: Neon is
    # serverless and its pooler can recycle/drop a long-idle or long-running
    # connection. Re-checking out a connection from the pool for each
    # machine lets pool_pre_ping validate (and transparently reconnect)
    # before each machine's writes, instead of one connection having to
    # survive the entire multi-minute, 31-machine operation uninterrupted.
    for i, csv_path in enumerate(files):
        machine = _machine_name_from_path(csv_path)
        df = pd.read_csv(csv_path)

        def _on_chunk(chunk_idx, total_chunks, _machine=machine, _i=i):
            progress.progress(
                _i / len(files),
                text=f"Saving {_machine} ({_i + 1}/{len(files)}) — chunk {chunk_idx}/{total_chunks}..."
            )

        progress.progress(i / len(files), text=f"Saving {machine} ({i + 1}/{len(files)})...")
        with engine.begin() as conn:
            total_rows += upsert_dataframe(conn, df, machine, on_chunk=_on_chunk)

    progress.progress(1.0, text="Done")

    return total_rows


def save_and_download_results(output_folder=SWITCHGEAR_OUTPUT_FOLDER):
    st.subheader("💾 Save & Download Results")

    files = _list_output_files(output_folder)

    if not files:
        st.info("No prediction results available yet. Run a prediction first.")
        return

    st.dataframe(_build_results_summary(output_folder), width='stretch')

    try:
        engine = get_engine()
    except Exception as e:
        st.error(f"Could not connect to database: {e}")
        return

    if st.button("🔌 Test Connection", key="sw_test_conn_btn"):
        with st.spinner("Connecting..."):
            try:
                version = test_connection(engine)
            except Exception as e:
                st.error(f"Connection failed: {e}")
            else:
                st.success(f"Connected ✅ — {version}")

    with st.expander("✍️ Manual Insert (single row)"):
        st.caption(
            "Writes one row directly — useful to check whether the database "
            "write itself works, separate from the full multi-row save below."
        )
        with st.form("sw_manual_insert_form"):
            machine = st.selectbox(
                "Machine", [f"A{i}" for i in range(1, 32)], key="sw_manual_machine"
            )
            date_val = st.date_input("Date", key="sw_manual_date")
            time_val = st.time_input("Time", key="sw_manual_time")
            health_index = st.number_input(
                "Health Index", value=1.0, step=1.0, key="sw_manual_health_index"
            )
            status = st.number_input(
                "Status", value=1, step=1, key="sw_manual_status"
            )
            manual_submitted = st.form_submit_button("Insert Row")

        if manual_submitted:
            row_df = pd.DataFrame([{
                "machine": machine,
                "datetime": datetime.combine(date_val, time_val),
                "date": date_val,
                "health_index": health_index,
                "status": int(status),
            }])
            with st.spinner("Inserting..."):
                try:
                    ensure_table_exists(engine)
                    with engine.begin() as conn:
                        n = upsert_dataframe(conn, row_df, machine)
                except Exception as e:
                    st.error(f"Manual insert failed: {e}")
                else:
                    st.success(f"Inserted/updated {n} row ✅")

    if st.button("💾 Save to Database", key="sw_save_btn"):
        with st.spinner("Saving results to database..."):
            try:
                total_rows = _upsert_output_folder(engine, output_folder)
            except Exception as e:
                st.error(f"Save failed: {e}")
            else:
                st.session_state.pop("sw_download_bytes", None)
                st.success(f"Saved/updated {total_rows} rows to the database ✅")

    with st.expander("⚠️ Reset Database"):
        reset_confirmed = st.checkbox(
            "I understand this will permanently delete all saved results.",
            key="sw_reset_confirm"
        )
        if st.button(
            "⚠️ Reset Database",
            key="sw_reset_btn",
            type="secondary",
            disabled=not reset_confirmed
        ):
            with st.spinner("Resetting database..."):
                try:
                    reset_table(engine)
                    total_rows = _upsert_output_folder(engine, output_folder)
                except Exception as e:
                    st.error(f"Reset failed: {e}")
                else:
                    st.session_state.pop("sw_download_bytes", None)
                    st.success(f"Database reset and reloaded with {total_rows} rows ✅")

    download_choice = st.radio(
        "Which file would you like to download?",
        ("Download last results.", "Download all database results."),
        key="sw_download_choice"
    )

    # Building the workbook (and, for the "all database results" option,
    # querying the whole table) is expensive — only do it when explicitly
    # requested, not on every rerun the tab happens to go through.
    if st.button("📦 Prepare Download", key="sw_prepare_download_btn"):
        with st.spinner("Preparing file..."):
            if download_choice == "Download last results.":
                wb = _build_workbook_from_output_folder(output_folder)
                file_name = "Switchgear_Last_Results.xlsx"
            else:
                try:
                    df = fetch_all_results(engine)
                except Exception as e:
                    st.error(f"Could not read database: {e}")
                    df = None

                if df is not None and df.empty:
                    st.warning("No saved results yet — click 'Save to Database' first.")
                    df = None

                wb = _build_workbook_from_dataframe(df) if df is not None else None
                file_name = "Switchgear_Database_Results.xlsx"

            if wb is not None:
                buffer = BytesIO()
                wb.save(buffer)
                st.session_state["sw_download_bytes"] = buffer.getvalue()
                st.session_state["sw_download_filename"] = file_name

    if st.session_state.get("sw_download_bytes"):
        st.download_button(
            label=f"Download {st.session_state['sw_download_filename']}",
            data=st.session_state["sw_download_bytes"],
            file_name=st.session_state["sw_download_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sw_download_btn"
        )

def func_main():

    st.header("Switchgear Health Index")

    raw_folder = "data/raw_data"
    file_count = count_files(raw_folder)

    # =====================================================
    # CONTROL PANEL
    # =====================================================
    with st.container(border=True):

        st.subheader("⚙️ Fleet Prediction Control Panel")

        col_status, col_run, col_clear = st.columns([2, 1, 1])

        # ===============================
        # DATA STATUS
        # ===============================
        with col_status:

            if file_count > 20:
                st.info(
                    f"Data available ({file_count} files). Ready for prediction."
                )
            else:
                st.warning(
                    f"⚠️ Not enough data. Found {file_count} files "
                    "(minimum required: 21)."
                )

        # ===============================
        # RUN PREDICTION BUTTON
        # ===============================
        with col_run:

            run_btn = st.button(
                "🚀 Run Prediction",
                disabled=(file_count <= 20)
            )

        # ===============================
        # CLEAR DATA BUTTON
        # ===============================
        with col_clear:

            clear_btn = st.button(
                "🗑 Clear output",
                type="secondary",
                help=(
                    "Delete all previous prediction results and output data. "
                    "Use this before running a new prediction."
                )
            )

    # =====================================================
    # RUN PREDICTION ACTION
    # =====================================================
    if run_btn:

        with st.spinner("Running fleet prediction..."):
            run_fleet_prediction()

        st.success("Prediction completed successfully ✅")

    # =====================================================
    # CLEAR DATA ACTION
    # =====================================================
    if clear_btn:

        with st.status("Resetting system...", expanded=True) as status:

            status.update(
                label="Removing previous results...",
                state="running"
            )
            clear_folder("data/output")

            status.update(
                label="System ready for new prediction 🚀",
                state="complete"
            )

    # =====================================================
    # DISPLAY RESULTS
    # =====================================================
    window_days = status_risk(path_output="data/output")

    status_explain(window_days)

    display_risk_by_case(
        path_raw="data/raw_data",
        path_output="data/output"
    )

    save_and_download_results()

