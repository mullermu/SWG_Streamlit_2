# pip install streamlit
# pip install numpy
# pip install joblib
# pip install openpyxl

# from tkinter import X
import streamlit as st
import pandas as pd
import os
import pandas as pd
import MC_status
import MC_Graph
import model_predict
import glob
import openpyxl 
import csv
from data_split import Data_Split

#-------------------------------------------
# addition v3
#-------------------------------------------
import tensorflow as tf
from source_code.util.data_processing import split_and_savefile
from source_code.util.app_switchgear_v2 import func_main
#-------------------------------------------

# from streamlitapp.Data_Input import data_input

st.set_page_config(
    page_title="Swtich Gear Prediction Main Page",
    layout="wide"
)
def get_csv(df):
    
    csv = df.to_csv(index=False)
    
    return csv
def get_data1(dtfile):
    dtfile.to_csv (r'Temp/temp.csv', index = False, header=True)

def form_callback():
    st.write(st.session_state.my_option)
    st.write(st.session_state.my_checkbox)

def listmodel(path):
    # return os.urllib(path)
    
    return os.listdir(path)

def listmachine(path):
    return os.listdir(path)



# =========================================================
# CLEAR FOLDER FUNCTION
# =========================================================
def clear_folder(folder_path):
    """Delete all files in a folder"""
    if os.path.exists(folder_path):
        for f in os.listdir(folder_path):
            file_path = os.path.join(folder_path, f)
            if os.path.isfile(file_path):
                os.remove(file_path)



def st_header(data):

    st.title("Switch Gear Status Classification App")

    output_status = False
    selected_model = None

    # =====================================================
    # INIT SESSION STATE
    # =====================================================
    if "upload_done" not in st.session_state:
        st.session_state.upload_done = False

    if "selected_model" not in st.session_state:
        st.session_state.selected_model = None

    if "uploader_key" not in st.session_state:
        st.session_state.uploader_key = 0

    if "start_upload" not in st.session_state:
        st.session_state.start_upload = False

    # =====================================================
    # UI CONTAINER
    # =====================================================
    with st.container():

        spacer1, main_col, spacer2 = st.columns([1, 10, 1])

        with main_col:

            # -------------------------------------------------
            # FILE UPLOADERS
            # -------------------------------------------------
            uploaded_file1 = st.file_uploader(
                "Choose A1-A16 file",
                type=["xlsx"],
                key=f"file1_{st.session_state.uploader_key}"
            )

            uploaded_file2 = st.file_uploader(
                "Choose A18-A31 file",
                type=["xlsx"],
                key=f"file2_{st.session_state.uploader_key}"
            )

            # =================================================
            # BUTTONS
            # =================================================
            col_upload, col_reset = st.columns(2)

            # Upload button
            with col_upload:
                if st.button(
                    "Upload Files",
                    disabled=not (uploaded_file1 and uploaded_file2)
                ):
                    st.session_state.start_upload = True

            # Clear button
            with col_reset:
                if st.button("🔄 Clear Status"):

                    st.session_state.start_upload = False
                    st.rerun()

                   
            # =====================================================
            # PROCESSING STATUS (FULL WIDTH)
            # =====================================================
            if st.session_state.start_upload:

                with st.status("Processing files...", expanded=True) as status:

                    v0_ok = False
                    v3_ok = False

                    progress_text = st.empty()   

                    # =================================================
                    # STEP 1 — MODEL v0 (FORMAT CHECK)
                    # =================================================
                    status.update(
                        label="Step 1/3 — Checking Model v0 format...",
                        state="running"
                    )

                    try:
                        # ---------- FILE 1 ----------
                        progress_text.info("📂 Checking files... (1/2)")

                        Data_Split(uploaded_file1).split()

                        # ---------- FILE 2 ----------
                        progress_text.info("📂 Checking files... (2/2)")

                        Data_Split(uploaded_file2).split()

                        v0_ok = True
                        progress_text.success("✅ Model v0 format detected")

                    except Exception:
                        v0_ok = False
                        progress_text.error("❌ Model v0 not supported")

                    # =================================================
                    # STEP 2 — model v3 (FULL PROCESSING)
                    # =================================================
                    status.update(
                        label="Step 2/3 — Checking model v3 format...",
                        state="running"
                    )

                    try:
                        # ---------- FILE 1 ----------
                        progress_text.info("📂 Processing files... (1/2)")

                        split_and_savefile(uploaded_file1).split(
                            status=status,
                            progress_text=progress_text  
                        )

                        # ---------- FILE 2 (ONLY IF FILE 1 OK) ----------
                        progress_text.info("📂 Processing files... (2/2)")

                        split_and_savefile(uploaded_file2).split(
                            status=status,
                            progress_text=progress_text
                        )

                        v3_ok = True
                        progress_text.success("✅ model v3 format detected")

                    except Exception:
                        v3_ok = False
                        progress_text.error("❌ model v3 not supported")

                    # =================================================
                    # STEP 3 — FINAL DECISION
                    # =================================================
                    status.update(
                        label="Step 3/3 — Validating dataset...",
                        state="running"
                    )

                    if not v0_ok and not v3_ok:

                        st.error("⛔ Dataset not supported by any model")
                        st.warning("Please check file format and upload again")

                        status.update(state="complete")
                        st.session_state.start_upload = False
                        return False, None

                    elif v0_ok and not v3_ok:

                        st.success("✅ Compatible with Model v0")
                        st.info("👉 Go to 'Select Model' section to continue")
                        selected_model = "v0"

                    elif v3_ok and not v0_ok:

                        st.success("✅ Compatible with model v3")
                        st.info("👉 Open 'Switchgear Health Index' tab")
                        selected_model = "v3"

                    else:

                        st.success("✅ Compatible with BOTH models")

                        st.warning(
                            "👉 Please choose model:\n"
                            "• Model v0 → 'Select Model'\n"
                            "• model v3 → 'Switchgear Health Index'"
                        )

                        selected_model = "v3"

                    st.info(f"➡️ Selected model (default): {selected_model}")

                    # =================================================
                    # COMPLETE
                    # =================================================
                    status.update(
                        label="Files processed successfully ✅",
                        state="complete"
                    )

            


        # SAVE STATE
        st.session_state.upload_done = True
        st.session_state.selected_model = selected_model
        st.session_state.start_upload = False
        output_status = True

    return output_status, selected_model


def st_body():
    
    # url = 'https://raw.githubusercontent.com/mullermu/SWG_Streamlit/tree/main/streamlitapp/model/'
    lstmodel = listmodel('model/')
    # st.write(lstmodel)
    tmp = [i.split('.')[0] for i in lstmodel]
    col1, col2, col3 = st.columns([1,10,1])
    with col2 :
        with st.form(key='my_form'):
            option = st.selectbox('Select Model (Model v0):',tmp,key="my_option")
            submitted = st.form_submit_button('selected model and predict')
            if submitted:
                st.write('You selected model: {}'.format(str(option)))
            return lstmodel[tmp.index(option)], submitted
                
    # import requests
    # from bs4 import BeautifulSoup

    # # URL on the Github where the csv files are stored
    # github_url = 'https://https://github.com/mullermu/SWG_Streamlit_2/tree/master/model'  # change USERNAME, REPOSITORY and FOLDER with actual name

    # result = requests.get(github_url)

    # soup = BeautifulSoup(database.text, 'html.parser')
    # csvfiles = soup.find_all(title=result.compile("\.csv$"))

    # filename = [ ]
    # for i in csvfiles:
    #         filename.append(i.extract().get_text())
        model_name = select_model()

def st_result(clf, df):
    if clf is not None:
        df = None
        model_predict.get_predict_result.getResult(clf, df, False)
        # model = joblib.load(os.path.join("../model/",clf))
        # z = model.predict(X)
        # res = pd.concat([data,pd.DataFrame(z,columns=['Status'])],axis=1)
        # col1, col2, col3 = st.columns([1,1,1])
        # with col2 :
        #     st.download_button("Download Classification File",get_csv(res),"../output/result_app.csv")
def create_last_results():
    files = [os.path.split(filename) for filename in glob.glob("output/Predicted Results/*.csv")]
    # write_only streams rows straight to a temp file instead of keeping
    # every cell as an in-memory object across all sheets, which matters
    # here: ~30 sheets x thousands of rows in normal mode can be a large
    # memory spike on a memory-constrained deployment.
    wb = openpyxl.Workbook(write_only=True)
    progress = st.progress(0.0, text="Preparing last results...")
    for i, (f_path, f_name) in enumerate(files):
        (f_short_name, f_extension) = os.path.splitext(f_name)
        progress.progress(i / len(files), text=f"Preparing last results — {f_short_name} ({i + 1}/{len(files)})...")
        with open(os.path.join(f_path, f_name)) as f_input:
            # st.write(f_short_name)
            ws = wb.create_sheet(title=os.path.basename(f_short_name))
            for row in csv.reader(f_input):
                ws.append(row)
    wb.save('output/Predicted Results/Last_Results.xlsx')
    progress.progress(1.0, text="Last results ready ✅")
def download_results():
        
    files = [os.path.split(filename) for filename in glob.glob("output/Predicted Results/*.csv")]
    st.write("Data Predicted")

    saveToDB = st.radio(
    "Do you want to save data to database?",
    ('Yes, This is new data.', 'No, Data was the same as the old one.','Replace all with the new one.'))
    saveBT = st.button('Save to Database')
    if saveToDB == 'Yes, This is new data.' and saveBT == True:
        with st.spinner("Saving to database..."):
            wb = openpyxl.load_workbook('output/Predicted Results/Results.xlsx')
            progress = st.progress(0.0, text="Saving...")
            for i, (f_path, f_name) in enumerate(files):
                (f_short_name, f_extension) = os.path.splitext(f_name)
                progress.progress(i / len(files), text=f"Saving {f_short_name} ({i + 1}/{len(files)})...")
                with open(os.path.join(f_path, f_name)) as f_input:
                    csv_reader = csv.reader(f_input)
                    first_line = next(csv_reader)
                    if f_short_name not in wb.sheetnames:
                        ws = wb.create_sheet(title=f_short_name)
                        ws.append(first_line)
                    else:
                        ws = wb[f_short_name]
                    for row in csv_reader:
                        if row != first_line:
                            ws.append(row)
            progress.progress(1.0, text="Writing workbook to disk...")
            wb.save('output/Predicted Results/Results.xlsx')
            create_last_results()
        st.success("Added to Database ✅")

    elif saveToDB == 'No, Data was the same as the old one.' and saveBT == True:
        with st.spinner("Preparing last results..."):
            create_last_results()
        st.info("Nothing saved")
    elif saveToDB == 'Replace all with the new one.' and saveBT == True:
        with st.spinner("Replacing database..."):
            wb = openpyxl.Workbook(write_only=True)
            progress = st.progress(0.0, text="Saving...")
            for i, (f_path, f_name) in enumerate(files):
                (f_short_name, f_extension) = os.path.splitext(f_name)
                progress.progress(i / len(files), text=f"Saving {f_short_name} ({i + 1}/{len(files)})...")
                with open(os.path.join(f_path, f_name)) as f_input:
                    # st.write(f_short_name)
                    ws = wb.create_sheet(title=os.path.basename(f_short_name))
                    for row in csv.reader(f_input):
                        ws.append(row)
            progress.progress(1.0, text="Writing workbook to disk...")
            wb.save('output/Predicted Results/Results.xlsx')
            create_last_results()
        st.success("Replaced database ✅")
    
    downloadResults = st.radio(
    "Which file do you would like to download?",
    ('Download last results.', 'Download all database results.'))
    if downloadResults == 'Download last results.':
        filePath = 'output/Predicted Results/Last_Results.xlsx'
        fileName = 'Last_Results.xlsx'
    elif downloadResults == 'Download all database results.':
        filePath = 'output/Predicted Results/Results.xlsx'
        fileName = 'Database_Results.xlsx'

    with open(filePath, 'rb') as my_file:
        st.download_button(label = 'Download Results', data = my_file.read(), file_name = fileName, mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def status_body():
    status = MC_status.MC_status.st_status()
    return status

def summary_chart():
    st.header("Summary Chart")
    chart_data = pd.DataFrame(
    MC_status.MC_status.sum_status(),
    columns=['Summary'],
    index=['Stage A', 'Stage B', 'Stage C','Stage D'])
    st.bar_chart(chart_data)

def history_chart():
    MC_Graph.mc_graph.get_mc_graph()

def preview_uploaded_data_v0(folder="Temp"):
    files = sorted(glob.glob(f"{folder}/*.csv"))

    if not files:
        return

    machine_names = [os.path.splitext(os.path.basename(f))[0] for f in files]

    with st.expander("📄 Preview Uploaded Data", expanded=False):
        selected_machine = st.selectbox(
            "Select machine to preview",
            machine_names,
            key="v0_preview_machine"
        )

        try:
            df = pd.read_csv(os.path.join(folder, f"{selected_machine}.csv"))
        except Exception as e:
            st.error(f"Could not read {selected_machine}.csv: {e}")
            return

        st.caption(f"{len(df)} rows × {len(df.columns)} columns")
        st.dataframe(df, width='stretch')

def tab_model_v0(status):
    st.header("Model V0")

    preview_uploaded_data_v0()

    with st.container(border=True):
        st.subheader("⚙️ Model Selection & Prediction")
        clf, submitted = st_body()

    # Only run the (heavyweight: load a joblib model + predict over every
    # Temp/ file) prediction on the exact rerun the button was actually
    # clicked — clf itself is never None (the form always has some model
    # selected), so gating on that alone would re-run this on every single
    # interaction anywhere in the app.
    if submitted and clf is not None and status is True:
        st_result(clf, None)

    st.subheader("📊 Fleet Status Summary")
    status_body()
    summary_chart()

    # Show Save & Download based on whether prediction output actually
    # exists on disk (persists across reruns), not the one-shot `submitted`
    # flag — otherwise interacting with this section's own radio/save
    # button would make the whole section disappear on the next rerun.
    has_predicted_output = bool(glob.glob("output/Predicted Results/A*.csv"))
    if has_predicted_output:
        st.subheader("💾 Save & Download Results")
        download_results()
    else:
        st.info("Please upload files and select a prediction model, then click 'selected model and predict'.")

    st.subheader("📈 History Chart")
    history_chart()


def main():
    with st.sidebar:
        status = None
        status, model = st_header(status)

    tab_v3, tab_v0 = st.tabs(["Switchgear Health Index (Model V3)", "Model V0"])

    with tab_v3:
        #-------------------------------------------
        # addition v3
        #-------------------------------------------
        func_main()

    with tab_v0:
        tab_model_v0(status)



main()
